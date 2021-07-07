#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jun 29 17:00:25 2021

@author: samuelsolomon
"""


# Basic Modules
import os
import sys
import numpy as np
# Read/Write to Excel
import csv
import pyexcel
import openpyxl as xl
# Plotting
import matplotlib.pyplot as plt



class dataProcessing:        
        
    def convertToXLSX(self, excelFile):
        """
        Converts .xls Files to .xlsx Files That OpenPyxl Can Read
        If the File is Already a .xlsx Files, Do Nothing
        If the File is Neither a .xls Nor .xlsx, it Exits the Program
        """
        # Check That the Current Extension is .xls or .xlsx
        _, extension = os.path.splitext(excelFile)
        # If the Extension is .xlsx, the File is Ready; Do Nothing
        if extension == '.xlsx':
            return excelFile
        # If the Extension is Not .xls/.xlsx, Then the Data is in the Wrong Format; Exit Program
        if extension not in ['.xls', '.xlsx']:
            print("Cannot Convert File to .xlsx")
            sys.exit()
        
        # Create Output File Directory to Save Data ONLY If None Exists
        newExcelFolder = os.path.dirname(excelFile) + "/Excel Files/"
        os.makedirs(newExcelFolder, exist_ok = True)
        
        # Convert '.xls' to '.xlsx'
        filename = os.path.basename(excelFile)
        newExcelFile = newExcelFolder + filename + "x"
        pyexcel.save_as(file_name = excelFile, dest_file_name = newExcelFile, logfile=open(os.devnull, 'w'))
        
        # Save New Excel name
        return newExcelFile
    
    def txt2csv(self, txtFile, csvFile, csvDelimiter = ",", overwriteCSV = False):
        # Check to see if csv conversion alreayd happened
        if not os.path.isfile(csvFile) or overwriteCSV:
            with open(txtFile, "r") as inputData:
                in_reader = csv.reader(inputData, delimiter = csvDelimiter)
                with open(csvFile, 'w', newline='') as out_csv:
                    out_writer = csv.writer(out_csv)
                    for row in in_reader:
                        out_writer.writerow(row)
    
    def convertToExcel(self, inputFile, excelFile, excelDelimiter = ",", overwriteXL = False, testSheetNum = 0):
        # If the File is Not Already Converted: Convert the CSV to XLSX
        if not os.path.isfile(excelFile) or overwriteXL:
            # Make Excel WorkBook
            xlWorkbook = xl.Workbook()
            xlWorksheet = xlWorkbook.active
            # Write the Data from the CSV File to the Excel WorkBook
            with open(inputFile, 'w+') as newFile:
                reader = csv.reader(newFile, delimiter = excelDelimiter)
                for row in reader:
                    xlWorksheet.append(row)
            # Save as New Excel File
            xlWorkbook.save(excelFile)
        # Else Load the GSR Data from the Excel File
        else:
            # Load the GSR Data from the Excel File
            xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
            xlWorksheet = xlWorkbook.worksheets[testSheetNum]
        
        # Return Excel Sheet
        return xlWorkbook, xlWorksheet
    

class processData(dataProcessing):
    
    def extractCosmolData(self, xlWorksheet, yVal = 25, zCol = 5):
        
        # -------------------------------------------------------------------#
        # ----------------------- Extract Run Info --------------------------#
        
        x = []; z = []; concentrations = []
        # Loop Through the Info Section and Extract the Needed Run Info from Excel
        rowGenerator = xlWorksheet.rows
        for cell in rowGenerator:
            
            if cell[1].value == yVal:
                x.append(cell[0].value)
                z.append(cell[2].value)
                concentrations.append(cell[zCol].value)
        
        return x, z, concentrations
    
    def getData(self, excelFile, testSheetNum = 0):
        """
        --------------------------------------------------------------------------
        Input Variable Definitions:
            excelFile: The Path to the Excel File Containing the Data
            testSheetNum: An Integer Representing the Excel Worksheet (0-indexed) Order.
        --------------------------------------------------------------------------
        """
        # Check if File Exists
        if not os.path.exists(excelFile):
            print("The following Input File Does Not Exist:", excelFile)
            sys.exit()
            
        # Convert TXT and CSV Files to XLSX
        if excelFile.endswith(".txt") or excelFile.endswith(".csv") or excelFile.endswith(".numbers"):
            # Extract Filename Information
            oldFileExtension = os.path.basename(excelFile)
            filename = os.path.splitext(oldFileExtension)[0]
            newFilePath = os.path.dirname(excelFile) + "/Excel Files/"
            # Make Output Folder Directory if Not Already Created
            os.makedirs(newFilePath, exist_ok = True)

            # Convert CSV or TXT to XLSX
            excelFile = newFilePath + filename + ".xlsx"
            xlWorkbook, xlWorksheet = self.convertToExcel(excelFile, excelFile, excelDelimiter = ",", overwriteXL = False, testSheetNum = testSheetNum)
        # If the File is Already an Excel File, Just Load the File
        elif excelFile.endswith(".xlsx"):
            # Load the GSR Data from the Excel File
            xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
            xlWorksheet = xlWorkbook.worksheets[testSheetNum]
        else:
            print("The Following File is Neither CSV, TXT, Nor XLSX:", excelFile)
            sys.exit()
        print("Extracting Data from the Excel File:", excelFile)
        
        # Extract Time and Current Data from the File
        xPoints, zPoints, concentrations = self.extractCosmolData(xlWorksheet, yVal = 25)
        
        xlWorkbook.close()
        # Finished Data Collection: Close Workbook and Return Data to User
        print("Done Collecting Data");
        return np.array(xPoints), np.array(zPoints), np.array(concentrations)


if __name__ == "__main__":
    
    cosmolFile = './Input Data/diffusion4.xlsx'
    x, y, z = processData().getData(cosmolFile)
    
    from scipy import interpolate
    import numpy as np
    from mpl_toolkits import mplot3d
    
    f = interpolate.interp2d(x, y, z, kind='quintic')
    xSamples = np.arange(-115,115,20)
    ySamples = np.arange(-115,115,20)
    zSamples = f(xSamples, ySamples)
    
    xx,yy=np.meshgrid(xSamples, ySamples)
    
    fig = plt.figure()
    ax = plt.axes(projection='3d');
    #ax.scatter3D(xSamples, ySamples, zSamples.ravel(), c=zSamples.ravel(), cmap='Greens');
    ax.scatter3D(xx.ravel(), yy.ravel(), zSamples.ravel(), c=zSamples.ravel(), cmap='Greens');
    
    plt.plot(x, z, 'ro-', xSamples, zSamples, 'b-')
    plt.show()
    
    
    